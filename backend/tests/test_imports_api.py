import hashlib
import json
from collections.abc import AsyncIterator
from io import BytesIO
from urllib.parse import urlsplit, urlunsplit

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.api.dependencies import get_import_document_fetcher
from app.core.database import get_session
from app.core.security import password_service
from app.domain.network import OutboundNetworkPolicy
from app.importers.sources import (
    FetchedImportDocument,
    ImportDocumentOption,
    ImportUrlDiscovery,
)
from app.main import app
from app.models import Base
from app.models.access import User

ADMIN_EMAIL = "import-admin@example.com"
ADMIN_PASSWORD = "import-password-123!"


class StubImportDocumentFetcher:
    def __init__(
        self,
        documents: dict[str, bytes],
        discoveries: dict[str, list[tuple[str, str]]] | None = None,
    ) -> None:
        self.documents = documents
        self.discoveries = discoveries or {}
        self.requests: list[tuple[str, int]] = []

    async def discover(
        self,
        *,
        url: str,
        network_policy: OutboundNetworkPolicy,
        maximum_bytes: int,
    ) -> ImportUrlDiscovery:
        del network_policy
        self.requests.append((url, maximum_bytes))
        options = self.discoveries.get(url, [("openapi.json", url)])
        return ImportUrlDiscovery(
            source_url=_safe_url(url),
            source_kind="swagger_ui" if url in self.discoveries else "document",
            documents=tuple(_stub_option(name, document_url) for name, document_url in options),
        )

    async def fetch(
        self,
        *,
        url: str,
        network_policy: OutboundNetworkPolicy,
        maximum_bytes: int,
        document_id: str | None = None,
    ) -> FetchedImportDocument:
        del network_policy
        self.requests.append((url, maximum_bytes))
        options = self.discoveries.get(url, [("openapi.json", url)])
        selected = next(
            (
                _stub_option(name, document_url)
                for name, document_url in options
                if document_id is None or _stub_id(document_url) == document_id
            ),
            None,
        )
        if selected is None:
            raise AssertionError("Unknown stub document selection")
        hostname = urlsplit(selected.url).hostname or "remote"
        return FetchedImportDocument(
            content=self.documents[selected.url],
            source_page_url=_safe_url(url),
            resolved_url=selected.url,
            source_name=f"{hostname}/{selected.name}",
            document_id=selected.id,
            discovered_from_page=url in self.discoveries,
        )


def _stub_id(url: str) -> str:
    return hashlib.sha256(url.encode()).hexdigest()


def _safe_url(url: str) -> str:
    parsed = urlsplit(url)
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", ""))


def _stub_option(name: str, url: str) -> ImportDocumentOption:
    return ImportDocumentOption(
        id=_stub_id(url),
        name=name,
        url=url,
        display_url=_safe_url(url),
    )


@pytest.fixture
async def import_client() -> AsyncIterator[AsyncClient]:
    test_engine = create_async_engine(
        "sqlite+aiosqlite://",
        poolclass=StaticPool,
        connect_args={"check_same_thread": False},
    )
    session_maker = async_sessionmaker(test_engine, expire_on_commit=False)
    async with test_engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    async with session_maker() as session:
        session.add(
            User(
                email=ADMIN_EMAIL,
                display_name="Import administrator",
                password_hash=password_service.hash(ADMIN_PASSWORD),
                is_active=True,
                is_system_admin=True,
                requires_password_change=False,
            )
        )
        await session.commit()

    async def override_session() -> AsyncIterator[AsyncSession]:
        async with session_maker() as session:
            yield session

    app.dependency_overrides[get_session] = override_session
    async with AsyncClient(
        transport=ASGITransport(app=app, raise_app_exceptions=False),
        base_url="http://test",
    ) as client:
        yield client
    app.dependency_overrides.clear()
    await test_engine.dispose()


@pytest.mark.asyncio
async def test_reimport_produces_diff_without_duplicate_definitions(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    first_document = _openapi_document(
        {
            "/users": {"get": {"summary": "List users"}},
            "/orders": {"post": {"summary": "Create order"}},
        }
    )

    first = await _upload_document(import_client, headers, project_id, first_document)
    assert first.status_code == 201, first.text
    assert first.json()["added"] == 2
    assert first.json()["changed"] == 0

    repeated = await _upload_document(import_client, headers, project_id, first_document)
    assert repeated.json()["unchanged"] == 2
    assert repeated.json()["added"] == 0

    changed_document = _openapi_document({"/users": {"get": {"summary": "List active users"}}})
    changed = await _upload_document(import_client, headers, project_id, changed_document)
    payload = changed.json()
    assert payload["changed"] == 1
    assert payload["deleted"] == 1
    assert {item["change"] for item in payload["results"]} == {"changed", "deleted"}
    deleted = next(item for item in payload["results"] if item["change"] == "deleted")
    assert deleted["method"] == "POST"
    assert deleted["path"] == "/orders"

    definitions = await import_client.get(
        f"/api/v1/projects/{project_id}/apis",
        headers=headers,
        params={"page": 1, "page_size": 100},
    )
    assert definitions.json()["total"] == 2
    users = next(
        item for item in definitions.json()["items"] if item["name"] == "List active users"
    )
    assert users["current_version"] == 2

    history = await import_client.get(f"/api/v1/projects/{project_id}/imports", headers=headers)
    assert history.status_code == 200
    assert history.json()["total"] == 3


@pytest.mark.asyncio
async def test_openapi_import_persists_complete_canonical_contract(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    document = json.dumps(
        {
            "openapi": "3.0.3",
            "info": {"title": "Orders", "version": "5.0.0"},
            "components": {"securitySchemes": {"bearerAuth": {"type": "http", "scheme": "bearer"}}},
            "paths": {
                "/orders/{tenantId}": {
                    "post": {
                        "operationId": "updateOrder",
                        "security": [{"bearerAuth": []}],
                        "parameters": [
                            {
                                "name": "tenantId",
                                "in": "path",
                                "required": True,
                                "schema": {"type": "string", "format": "uuid"},
                            },
                            {
                                "name": "dry_run",
                                "in": "query",
                                "schema": {"type": "boolean"},
                            },
                            {
                                "name": "X-Tenant-Id",
                                "in": "header",
                                "required": True,
                                "schema": {"type": "string", "minLength": 1},
                            },
                        ],
                        "requestBody": {
                            "required": True,
                            "content": {
                                "application/json": {
                                    "schema": {
                                        "type": "object",
                                        "required": ["quantity", "type", "profile"],
                                        "properties": {
                                            "quantity": {
                                                "type": "integer",
                                                "minimum": 1,
                                                "maximum": 999,
                                            },
                                            "type": {
                                                "type": "string",
                                                "enum": ["NORMAL", "PRIORITY"],
                                            },
                                            "remark": {"type": "string", "maxLength": 20},
                                            "profile": {
                                                "type": "object",
                                                "required": ["display_name"],
                                                "properties": {
                                                    "display_name": {
                                                        "type": "string",
                                                        "minLength": 1,
                                                    }
                                                },
                                            },
                                        },
                                    }
                                }
                            },
                        },
                        "responses": {
                            "200": {
                                "description": "updated",
                                "content": {
                                    "application/json": {
                                        "schema": {
                                            "type": "object",
                                            "required": ["id"],
                                            "properties": {"id": {"type": "string"}},
                                        }
                                    }
                                },
                            },
                            "400": {"description": "invalid"},
                            "401": {"description": "unauthorized"},
                        },
                    }
                }
            },
        }
    ).encode()

    imported = await _upload_document(import_client, headers, project_id, document)
    assert imported.status_code == 201, imported.text
    definition_id = imported.json()["results"][0]["definition_id"]
    detail = await import_client.get(
        f"/api/v1/projects/{project_id}/apis/{definition_id}", headers=headers
    )
    assert detail.status_code == 200, detail.text
    version = detail.json()["version"]

    assert version["contract_completeness"] == "complete"
    assert len(version["contract_fingerprint"]) == 64
    assert version["canonical_contract"]["parameters"][0]["location"] == "path"
    assert version["canonical_contract"]["responses"]["200"]["schema"]["required"] == ["id"]
    generated = await import_client.post(
        f"/api/v1/projects/{project_id}/test-engineering/generate",
        headers=headers,
        json={"api_definition_id": definition_id},
    )
    assert generated.status_code == 200, generated.text
    design = generated.json()["design"]
    mutations = [
        (mutation["location"], mutation["path"], mutation.get("value"))
        for scenario in design["scenarios"]
        for mutation in scenario["mutations"]
    ]
    assert ("body", "body.quantity", 1000) in mutations
    assert ("path", "path.tenantId", "not-a-uuid") in mutations
    assert any(
        location == "query" and path == "query.dry_run" and value == "invalid"
        for location, path, value in mutations
    )
    assert ("header", "header.X-Tenant-Id", None) in mutations
    assert ("body", "body.quantity", 0) in mutations
    assert ("body", "body.quantity", 1) in mutations
    assert ("body", "body.quantity", 999) in mutations
    assert ("body", "body.type", "__invalid__") in mutations
    assert ("body", "body.remark", "x" * 21) in mutations
    assert any(path == "body.profile.display_name" for _location, path, _value in mutations)
    assert ("auth", "auth", None) in mutations
    happy = next(item for item in design["scenarios"] if item["kind"] == "happy_path")
    assert happy["request"]["body"]["profile"]["display_name"]
    assert any(oracle["kind"] == "schema" for oracle in design["oracles"])


@pytest.mark.asyncio
async def test_import_rejects_invalid_and_duplicate_operations(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    invalid = await _upload_document(import_client, headers, project_id, b"[]")
    assert invalid.status_code == 422
    assert invalid.json()["error"]["code"] == "IMPORT_INVALID"

    duplicate = _openapi_document(
        {
            "/users": {"get": {"summary": "One"}},
            "//users": {"get": {"summary": "Duplicate"}},
        }
    )
    response = await _upload_document(import_client, headers, project_id, duplicate)
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "IMPORT_DUPLICATE_OPERATION"


@pytest.mark.asyncio
async def test_import_preview_selective_merge_and_explicit_deactivation(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    original = _openapi_document(
        {
            "/users": {"get": {"summary": "List users"}},
            "/orders": {"post": {"summary": "Create order"}},
        }
    )
    assert (await _upload_document(import_client, headers, project_id, original)).status_code == 201

    changed = _openapi_document(
        {
            "/users": {"get": {"summary": "List active users"}},
            "/products": {"get": {"summary": "List products"}},
        }
    )
    preview = await _preview_document(import_client, headers, project_id, changed)
    assert preview.status_code == 201, preview.text
    diff = preview.json()
    assert diff["status"] == "preview"
    assert {item["change"] for item in diff["results"]} == {
        "added",
        "changed",
        "deleted",
    }
    assert (
        next(item for item in diff["results"] if item["change"] == "added")["definition_id"] is None
    )

    selected = {
        item["import_key"] for item in diff["results"] if item["change"] in {"added", "changed"}
    }
    merged = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/{diff['id']}/merge",
        headers=headers,
        json={"selected_keys": sorted(selected)},
    )
    assert merged.status_code == 200, merged.text
    assert merged.json()["status"] == "applied"
    assert set(merged.json()["applied_keys"]) == selected

    definitions = await import_client.get(
        f"/api/v1/projects/{project_id}/apis",
        headers=headers,
        params={"page": 1, "page_size": 100},
    )
    assert definitions.json()["total"] == 3
    assert {item["name"] for item in definitions.json()["items"]} == {
        "List active users",
        "Create order",
        "List products",
    }

    deletion_preview = await _preview_document(import_client, headers, project_id, changed)
    deletion = next(
        item for item in deletion_preview.json()["results"] if item["change"] == "deleted"
    )
    deactivated = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/{deletion_preview.json()['id']}/merge",
        headers=headers,
        json={"selected_keys": [deletion["import_key"]]},
    )
    assert deactivated.status_code == 200
    assert deactivated.json()["applied_keys"] == [deletion["import_key"]]

    active = await import_client.get(
        f"/api/v1/projects/{project_id}/apis",
        headers=headers,
        params={"page": 1, "page_size": 100},
    )
    assert active.json()["total"] == 2

    repeated = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/{deletion_preview.json()['id']}/merge",
        headers=headers,
        json={"selected_keys": [deletion["import_key"]]},
    )
    assert repeated.status_code == 200
    different = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/{deletion_preview.json()['id']}/merge",
        headers=headers,
        json={"selected_keys": []},
    )
    assert different.status_code == 409
    assert different.json()["error"]["code"] == "IMPORT_ALREADY_APPLIED"


@pytest.mark.asyncio
async def test_url_import_previews_and_keeps_same_named_sources_separate(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    service_a = "https://service-a.example/openapi.json?version=1"
    service_b = "https://service-b.example/openapi.json"
    fetcher = StubImportDocumentFetcher(
        {
            service_a: _openapi_document(
                {
                    "/users": {"get": {"summary": "List users"}},
                    "/legacy": {"get": {"summary": "Legacy users"}},
                }
            ),
            service_b: _openapi_document({"/orders": {"get": {"summary": "List orders"}}}),
        }
    )
    app.dependency_overrides[get_import_document_fetcher] = lambda: fetcher

    preview_a = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/url/preview",
        headers=headers,
        json={"url": service_a, "source_type": "auto"},
    )
    assert preview_a.status_code == 201, preview_a.text
    payload_a = preview_a.json()
    assert payload_a["source_kind"] == "url"
    assert payload_a["source_name"] == "service-a.example/openapi.json"
    assert payload_a["source_url"] == "https://service-a.example/openapi.json"
    assert payload_a["document_url"] == "https://service-a.example/openapi.json"
    assert payload_a["source_key"].startswith("url:")
    assert "version=1" not in str(payload_a)
    selected_a = [item["import_key"] for item in payload_a["results"]]
    merged_a = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/{payload_a['id']}/merge",
        headers=headers,
        json={"selected_keys": selected_a},
    )
    assert merged_a.status_code == 200, merged_a.text

    preview_b = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/url/preview",
        headers=headers,
        json={"url": service_b},
    )
    assert preview_b.status_code == 201, preview_b.text
    payload_b = preview_b.json()
    assert payload_b["source_name"] == "service-b.example/openapi.json"
    assert payload_b["source_key"] != payload_a["source_key"]
    assert payload_b["added"] == 1
    assert payload_b["deleted"] == 0
    assert fetcher.requests[0][0] == service_a
    assert fetcher.requests[0][1] == 50 * 1024 * 1024


@pytest.mark.asyncio
async def test_swagger_ui_discovery_selects_group_and_redacts_document_query(
    import_client: AsyncClient,
) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    page_url = "https://service.example/swagger-ui/index.html"
    users_url = "https://service.example/v3/api-docs/users?token=private"
    orders_url = "https://service.example/v3/api-docs/orders"
    fetcher = StubImportDocumentFetcher(
        {
            users_url: _openapi_document({"/users": {"get": {"summary": "List users"}}}),
            orders_url: _openapi_document({"/orders": {"get": {"summary": "List orders"}}}),
        },
        {page_url: [("用户服务", users_url), ("订单服务", orders_url)]},
    )
    app.dependency_overrides[get_import_document_fetcher] = lambda: fetcher

    discovery = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/url/discover",
        headers=headers,
        json={"url": page_url},
    )
    assert discovery.status_code == 200, discovery.text
    discovery_payload = discovery.json()
    assert discovery_payload == {
        "source_url": page_url,
        "source_kind": "swagger_ui",
        "documents": [
            {
                "id": _stub_id(users_url),
                "name": "用户服务",
                "url": "https://service.example/v3/api-docs/users",
            },
            {
                "id": _stub_id(orders_url),
                "name": "订单服务",
                "url": orders_url,
            },
        ],
    }
    assert "private" not in discovery.text

    preview = await import_client.post(
        f"/api/v1/projects/{project_id}/imports/url/preview",
        headers=headers,
        json={
            "url": page_url,
            "source_type": "auto",
            "document_id": _stub_id(orders_url),
        },
    )
    assert preview.status_code == 201, preview.text
    payload = preview.json()
    assert payload["source_url"] == page_url
    assert payload["document_url"] == orders_url
    assert payload["source_name"] == "service.example/订单服务"
    assert payload["added"] == 1


@pytest.mark.asyncio
async def test_har_curl_bruno_excel_import_and_export(import_client: AsyncClient) -> None:
    headers = await _login_headers(import_client)
    project_id = await _create_project(import_client, headers)
    documents = [
        (
            "traffic.har",
            "auto",
            json.dumps(
                {
                    "log": {
                        "entries": [
                            {
                                "comment": "=2+2",
                                "request": {
                                    "method": "GET",
                                    "url": "https://api.example.com/har-users?page=1",
                                    "headers": [],
                                },
                            }
                        ]
                    }
                }
            ).encode(),
        ),
        (
            "request.curl",
            "curl",
            (
                b"curl -X POST 'https://api.example.com/curl-orders?token=live-token' "
                b"-H 'Authorization: Bearer live-auth' "
                b'--json \'{"password":"live-password","sku":"A1"}\''
            ),
        ),
        (
            "collection.bruno.json",
            "auto",
            json.dumps(
                {
                    "bruno": "FlowTest Collection",
                    "items": [
                        {
                            "name": "Bruno products",
                            "request": {
                                "method": "GET",
                                "url": "/bruno-products?active=true",
                                "headers": {"Accept": "application/json"},
                            },
                        }
                    ],
                }
            ).encode(),
        ),
        ("apis.xlsx", "excel", _excel_document()),
    ]
    for filename, source_type, content in documents:
        response = await _upload_document(
            import_client,
            headers,
            project_id,
            content,
            filename=filename,
            source_type=source_type,
        )
        assert response.status_code == 201, response.text
        assert response.json()["added"] == 1

    definitions = await import_client.get(
        f"/api/v1/projects/{project_id}/apis",
        headers=headers,
        params={"page": 1, "page_size": 100},
    )
    assert definitions.json()["total"] == 4
    curl_api = next(item for item in definitions.json()["items"] if item["name"] == "cURL import")
    detail = await import_client.get(
        f"/api/v1/projects/{project_id}/apis/{curl_api['id']}", headers=headers
    )
    assert detail.json()["version"]["headers"]["Authorization"].startswith("{{secret.")
    assert detail.json()["version"]["query_parameters"][0]["value"].startswith("{{secret.")
    assert detail.json()["version"]["body"]["password"].startswith("{{secret.")

    for export_format, suffix in (
        ("har", ".har"),
        ("curl", ".curl.txt"),
        ("bruno", ".bruno.json"),
        ("excel", ".xlsx"),
    ):
        exported = await import_client.get(
            f"/api/v1/projects/{project_id}/exports/apis",
            headers=headers,
            params={"format": export_format},
        )
        assert exported.status_code == 200, exported.text
        assert suffix in exported.headers["content-disposition"]
        assert exported.content
        if export_format == "excel":
            assert exported.content.startswith(b"PK")
            workbook = load_workbook(BytesIO(exported.content), data_only=False)
            sheet = workbook.active
            assert sheet is not None
            values = [str(cell.value) for row in sheet.iter_rows() for cell in row]
            assert "'=2+2" in values
            serialized = "\n".join(values)
            assert "live-token" not in serialized
            assert "live-password" not in serialized
            assert "live-auth" not in serialized
        if export_format == "curl":
            assert b"curl -X" in exported.content
        if export_format != "excel":
            assert b"live-token" not in exported.content
            assert b"live-password" not in exported.content
            assert b"live-auth" not in exported.content


async def _login_headers(client: AsyncClient) -> dict[str, str]:
    response = await client.post(
        "/api/v1/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}
    )
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


async def _create_project(client: AsyncClient, headers: dict[str, str]) -> str:
    response = await client.post(
        "/api/v1/projects",
        headers=headers,
        json={"name": "Imported APIs", "description": "Import verification"},
    )
    assert response.status_code == 201
    return str(response.json()["id"])


async def _upload_document(
    client: AsyncClient,
    headers: dict[str, str],
    project_id: str,
    content: bytes,
    *,
    filename: str = "sample.json",
    source_type: str = "auto",
):
    return await client.post(
        f"/api/v1/projects/{project_id}/imports",
        headers=headers,
        files={"document": (filename, content, "application/octet-stream")},
        data={"source_type": source_type},
    )


async def _preview_document(
    client: AsyncClient,
    headers: dict[str, str],
    project_id: str,
    content: bytes,
):
    return await client.post(
        f"/api/v1/projects/{project_id}/imports/preview",
        headers=headers,
        files={"document": ("sample.json", content, "application/json")},
        data={"source_type": "auto"},
    )


def _openapi_document(paths: dict[str, object]) -> bytes:
    return json.dumps(
        {
            "openapi": "3.0.3",
            "info": {"title": "Sample", "version": "1.0.0"},
            "paths": paths,
        }
    ).encode()


def _excel_document() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        [
            "name",
            "method",
            "path",
            "description",
            "query",
            "headers",
            "body",
            "auth_kind",
            "auth_config",
        ]
    )
    sheet.append(
        [
            "Excel inventory",
            "GET",
            "/excel-inventory",
            "Imported from Excel",
            '{"limit": 10}',
            '{"Accept": "application/json"}',
            "",
            "none",
            "{}",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
