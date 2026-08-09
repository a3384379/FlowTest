import json
from io import BytesIO

from openpyxl import load_workbook

from app.domain.api_assets import APIVersionSpec, AuthKind, BodyKind, HttpMethod, QueryParameterSpec
from app.importers.contracts import ImportedOperation, imported_value


class ExcelImportError(ValueError):
    """Raised when an Excel API document is malformed."""


def parse_excel(content: bytes) -> tuple[ImportedOperation, ...]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (OSError, ValueError, KeyError) as error:
        raise ExcelImportError("Excel 文件损坏或格式不受支持") from error
    sheet = workbook.active
    if sheet is None:
        raise ExcelImportError("Excel 文件没有可读取的工作表")
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise ExcelImportError("Excel 文件为空")
    headers = [str(value).strip().lower() if value is not None else "" for value in header_row]
    required = {"name", "method", "path"}
    if not required <= set(headers):
        raise ExcelImportError("Excel 表头必须包含 name、method、path")
    indexes = {name: headers.index(name) for name in headers if name}
    operations: list[ImportedOperation] = []
    for number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        operations.append(_operation(row, indexes, number))
    return tuple(operations)


def _operation(row: tuple[object, ...], indexes: dict[str, int], number: int) -> ImportedOperation:
    name = _cell(row, indexes, "name")
    method_value = _cell(row, indexes, "method").upper()
    path = _cell(row, indexes, "path")
    if not name or not path:
        raise ExcelImportError(f"Excel 第 {number} 行 name/path 不能为空")
    try:
        method = HttpMethod(method_value)
    except ValueError as error:
        raise ExcelImportError(f"Excel 第 {number} 行 HTTP 方法无效") from error
    query_values = _json_object(_cell(row, indexes, "query"), number, "query")
    header_values = _json_object(_cell(row, indexes, "headers"), number, "headers")
    auth_values = _json_object(_cell(row, indexes, "auth_config"), number, "auth_config")
    auth_kind_value = _cell(row, indexes, "auth_kind") or "none"
    try:
        auth_kind = AuthKind(auth_kind_value)
    except ValueError as error:
        raise ExcelImportError(f"Excel 第 {number} 行 auth_kind 无效") from error
    body_text = _cell(row, indexes, "body")
    body: object = None
    body_kind = BodyKind.NONE
    if body_text:
        body_kind = BodyKind.JSON
        try:
            body = json.loads(body_text)
        except json.JSONDecodeError as error:
            raise ExcelImportError(f"Excel 第 {number} 行 body 不是有效 JSON") from error
    return ImportedOperation(
        name=name[:200],
        description=_cell(row, indexes, "description")[:4000],
        request=APIVersionSpec(
            method=method,
            path=path,
            query_parameters=tuple(
                QueryParameterSpec(
                    name=key,
                    value=imported_value(key, str(value)),
                    enabled=True,
                )
                for key, value in query_values.items()
            ),
            headers={key: imported_value(key, str(value)) for key, value in header_values.items()},
            body_kind=body_kind,
            body=body,  # type: ignore[arg-type]
            auth_kind=auth_kind,
            auth_config={
                key: imported_value(key, str(value)) for key, value in auth_values.items()
            },
        ),
    )


def _cell(row: tuple[object, ...], indexes: dict[str, int], name: str) -> str:
    index = indexes.get(name)
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _json_object(value: str, number: int, field: str) -> dict[str, object]:
    if not value:
        return {}
    try:
        result = json.loads(value)
    except json.JSONDecodeError as error:
        raise ExcelImportError(f"Excel 第 {number} 行 {field} 不是有效 JSON") from error
    if not isinstance(result, dict) or not all(isinstance(key, str) for key in result):
        raise ExcelImportError(f"Excel 第 {number} 行 {field} 必须是 JSON 对象")
    return result
