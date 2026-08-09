import csv
import json
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import PurePath
from typing import cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import JsonValue

from app.engine.contracts import DatasetFormat

DATASET_ROW_LIMIT = 1_000
DATASET_COLUMN_LIMIT = 200


@dataclass(frozen=True, slots=True)
class DatasetParseError(ValueError):
    code: str
    message: str

    def __str__(self) -> str:
        return self.message


@dataclass(frozen=True, slots=True)
class ParsedDataset:
    format: DatasetFormat
    columns: tuple[str, ...]
    rows: tuple[dict[str, JsonValue], ...]


def parse_dataset(
    *,
    filename: str,
    content: bytes,
    requested_format: DatasetFormat,
    sheet_name: str | None = None,
) -> ParsedDataset:
    dataset_format = _resolve_format(filename, requested_format)
    try:
        if dataset_format is DatasetFormat.CSV:
            rows = _parse_csv(content)
        elif dataset_format is DatasetFormat.JSON:
            rows = _parse_json(content)
        else:
            rows = _parse_excel(content, sheet_name)
    except DatasetParseError:
        raise
    except (UnicodeDecodeError, json.JSONDecodeError, BadZipFile, InvalidFileException) as error:
        raise DatasetParseError("INVALID_DATASET", "数据集文件格式无效") from error
    _validate_rows(rows)
    return ParsedDataset(
        format=dataset_format,
        columns=tuple(rows[0]),
        rows=tuple(rows),
    )


def _resolve_format(filename: str, requested: DatasetFormat) -> DatasetFormat:
    if requested is not DatasetFormat.AUTO:
        return requested
    extension = PurePath(filename).suffix.lower()
    formats = {
        ".csv": DatasetFormat.CSV,
        ".json": DatasetFormat.JSON,
        ".xlsx": DatasetFormat.EXCEL,
    }
    resolved = formats.get(extension)
    if resolved is None:
        raise DatasetParseError(
            "UNSUPPORTED_DATASET_FORMAT",
            "数据集仅支持 CSV、JSON 和 XLSX",
        )
    return resolved


def _parse_csv(content: bytes) -> list[dict[str, JsonValue]]:
    table = list(csv.reader(StringIO(content.decode("utf-8-sig"))))
    if not table:
        return []
    headers = _headers(table[0])
    rows: list[dict[str, JsonValue]] = []
    for raw in table[1:]:
        if not any(value.strip() for value in raw):
            continue
        if len(raw) > len(headers):
            raise DatasetParseError("INCONSISTENT_DATASET", "数据集行包含未命名的额外字段")
        padded = [*raw, *([""] * max(0, len(headers) - len(raw)))]
        rows.append(dict(zip(headers, padded[: len(headers)], strict=True)))
    return rows


def _parse_json(content: bytes) -> list[dict[str, JsonValue]]:
    loaded: JsonValue = json.loads(content.decode("utf-8-sig"))
    if isinstance(loaded, dict) and isinstance(loaded.get("items"), list):
        loaded = loaded["items"]
    if not isinstance(loaded, list) or any(not isinstance(item, dict) for item in loaded):
        raise DatasetParseError("INVALID_DATASET", "JSON 数据集必须是对象数组")
    return [cast(dict[str, JsonValue], item) for item in loaded]


def _parse_excel(content: bytes, sheet_name: str | None) -> list[dict[str, JsonValue]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            raise DatasetParseError("DATASET_SHEET_NOT_FOUND", "指定的 Excel 工作表不存在")
        sheet = workbook[sheet_name] if sheet_name is not None else workbook.active
        if sheet is None:
            raise DatasetParseError("EMPTY_DATASET", "Excel 数据集没有可读取的工作表")
        values = sheet.iter_rows(values_only=True)
        first = next(values, None)
        if first is None:
            return []
        headers = _headers(first)
        rows: list[dict[str, JsonValue]] = []
        for raw in values:
            if not any(value is not None for value in raw):
                continue
            if len(raw) > len(headers) and any(value is not None for value in raw[len(headers) :]):
                raise DatasetParseError("INCONSISTENT_DATASET", "数据集行包含未命名的额外字段")
            padded = [*raw, *([None] * max(0, len(headers) - len(raw)))]
            rows.append(
                {
                    header: _excel_value(value)
                    for header, value in zip(headers, padded[: len(headers)], strict=True)
                }
            )
        return rows
    finally:
        workbook.close()


def _headers(values: tuple[object, ...] | list[str]) -> tuple[str, ...]:
    headers = tuple(str(value).strip() if value is not None else "" for value in values)
    if not headers or any(not header for header in headers):
        raise DatasetParseError("INVALID_DATASET_HEADERS", "数据集表头不能为空")
    if len(headers) != len(set(headers)):
        raise DatasetParseError("INVALID_DATASET_HEADERS", "数据集表头不能重复")
    if len(headers) > DATASET_COLUMN_LIMIT:
        raise DatasetParseError("DATASET_TOO_WIDE", "数据集最多包含 200 列")
    return headers


def _validate_rows(rows: list[dict[str, JsonValue]]) -> None:
    if not rows:
        raise DatasetParseError("EMPTY_DATASET", "数据集至少需要一行数据")
    if len(rows) > DATASET_ROW_LIMIT:
        raise DatasetParseError("DATASET_ROW_LIMIT", "数据集单次最多执行 1000 行")
    columns = tuple(rows[0])
    for index, row in enumerate(rows):
        if set(row) != set(columns):
            raise DatasetParseError("INCONSISTENT_DATASET", "数据集每行字段必须一致")
        rows[index] = {column: row[column] for column in columns}
    _headers(list(columns))


def _excel_value(value: object) -> JsonValue:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)
